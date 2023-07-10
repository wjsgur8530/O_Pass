#!/usr/bin/env python
from flask import Flask, flash, session, url_for, render_template, request, redirect, jsonify
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash # Password Hash
from flask_bcrypt import Bcrypt
from app import User, Visitor, Card, User_log, Year, Month, Day, Department, Rack, Privacy
import jinja2.exceptions
from config import create_app, db
from flask_login import login_user, login_required, logout_user, current_user
from flask_login import LoginManager
from datetime import datetime, date, time
from sqlalchemy import func, Integer, and_, or_
import qrcode
import mysql.connector
import openpyxl
import json
import pandas as pd
from datetime import datetime, timedelta
from tabulate import tabulate
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import re
from flask.helpers import get_flashed_messages
from db_connector import db_connector, account_list
app = create_app()
bcrypt = Bcrypt(app)

# Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = u"로그인 후에 서비스를 이용해주세요."
login_manager.login_message_category = "info"
@login_manager.user_loader
def load_user(id):
    return db.session.get(User, id)  # primary_key

#===================================================================================

@app.route('/')
@app.route('/index')
@login_required
def index():
    # 타임 스탬프
    today_weekday = datetime.now().weekday()
    weekdays = {0: "월요일", 1: "화요일", 2: "수요일", 3: "목요일", 4: "금요일", 5: "토요일", 6: "일요일"}
    weekday = weekdays.get(today_weekday, "")
    current_date = datetime.now().strftime('%Y년 %m월 %d일 ') + weekday
    current_time = datetime.now().strftime('\n%p %H:%M:%S')
    time = [ current_date, current_time]

    # 승인된 방문객 Sort_Desc
    approve_visitors = Visitor.query.filter_by(approve=1).order_by(Visitor.id.desc())

    # 출입 카드 목록
    card_list = Card.query.all()

    if approve_visitors:
        # 실시간 출입 방문객
        in_visitor = Visitor.query.filter_by(exit=0)
        in_visitor_card_none = Visitor.query.filter_by(exit=0, card_id=None)

        # 실시간 출입 방문객 카운팅
        in_visitor = in_visitor.count()
        in_visitor_card_none = in_visitor_card_none.count()
        in_visitor = in_visitor - in_visitor_card_none

        today = date.today()
        total_visitors = db.session.query(func.sum(Year.count)).scalar() # 총 방문객
        year = Year.query.filter_by(year=today.year).first() # 연간 방문객
        month = Month.query.filter_by(year=today.year, month=today.month).first() # 월간 방문객
        day = Day.query.filter_by(year=today.year, month=today.month, day=today.day).first() # 일간 방문객
        if year:
            yearly_visitor = year.count
            if month:
                monthly_visitor = month.count
                if day:
                    daily_visitor = day.count
                    visitor_count = [in_visitor, yearly_visitor, monthly_visitor, daily_visitor]
                else:
                    visitor_count = [in_visitor, yearly_visitor, monthly_visitor, 0]
            else:
                visitor_count = [in_visitor, yearly_visitor, 0, 0]
        else:
            visitor_count = [in_visitor, 0, 0, 0]
    
    # print('현재 로그인한 사용자: ' + str(current_user))
    return render_template('index.html', current_user=current_user, approve_visitors=approve_visitors, visitor_count=visitor_count, time=time, card_list=card_list, total_visitors=total_visitors)

#===================================================================================


#===================================================================================
admin_email, admin_password, admin2_email, admin2_password = account_list()
@app.route('/admin', methods=['GET', 'POST'])
def admin_page():
    if request.method == 'POST':
        admin = User.query.filter_by(username='관리자').first()
        if admin:
            flash('이미 관리자 계정이 존재합니다.')
            return redirect('admin')
        username = '관리자'
        email = admin_email
        hashed_password = bcrypt.generate_password_hash(admin_password)
        admin = User(username, email, hashed_password, 'Admin', 'M')
        db.session.add(admin)
        db.session.commit()
        flash("관리자 계정이 생성되었습니다.")
        return redirect('login')
    return render_template('admin.html')

@app.route('/admin_2', methods=['GET', 'POST'])
def admin2_page():
    if request.method == 'POST':
        admin = User.query.filter_by(username='상황실').first()
        if admin:
            flash('이미 상황실 계정이 존재합니다.')
            return redirect('admin')
        username = '상황실'
        email = admin2_email
        hashed_password = bcrypt.generate_password_hash(admin2_password)
        admin = User(username, email, hashed_password, 'Admin', 'S')
        db.session.add(admin)
        db.session.commit()
        flash("상황실 계정이 생성되었습니다.")
        return redirect('login')

#===================================================================================

#===================================================================================

# 차트 관리 페이지
@app.route('/charts', methods=['GET', 'POST'])
@login_required
def visualization_chart():
    today = date.today()
    total_visitors = db.session.query(func.sum(Year.count)).scalar() # 총 방문객
    year = Year.query.filter_by(year=today.year).first() # 연간 방문객
    month = Month.query.filter_by(year=today.year, month=today.month).first() # 월간 방문객
    day = Day.query.filter_by(year=today.year, month=today.month, day=today.day).first() # 일간 방문객
    if year:
        yearly_visitor = year.count
        if month:
            monthly_visitor = month.count
            if day:
                daily_visitor = day.count
                visitor_count = [yearly_visitor, monthly_visitor, daily_visitor]
            else:
                visitor_count = [yearly_visitor, monthly_visitor, 0]
        else:
            visitor_count = [yearly_visitor, 0, 0]
    else:
        visitor_count = [0, 0, 0]

    daily_visitors = Day.query.filter_by(month=datetime.now().strftime('%m')).all()
    monthly_visitors = Month.query.filter_by(year=datetime.now().strftime('%Y')).all()
    yearly_visitors = Year.query.all()

    # 일간 방문자 수
    day_count = []
    what_day = []
    what_month = []

    # 월간 방문자 수
    month_count = []
    what_month_2 = []

    # 연간 방문자 수
    what_year = []
    year_count = []
    
    for daily_visitor in daily_visitors:
        day_count.append(daily_visitor.count)
        what_day.append(daily_visitor.day)
        what_month.append(daily_visitor.month)
    daily = [what_month, what_day, day_count]

    for monthly_visitor in monthly_visitors:
        what_month_2.append(monthly_visitor.month)
        month_count.append(monthly_visitor.count)
    monthly = [what_month_2, month_count]

    for yearly_visitor in yearly_visitors:
        what_year.append(yearly_visitor.year)
        year_count.append(yearly_visitor.count)
    yearly = [what_year, year_count]
    print(daily, monthly, yearly)
    return render_template('charts.html', daily=daily, monthly=monthly, yearly=yearly, visitor_count=visitor_count, total_visitors=total_visitors)

#===================================================================================


#===================================================================================

# 방문객 관리 페이지
@app.route('/manage_visitors', methods=['GET', 'POST'])
@login_required
def manage_visitors():
    if request.method == 'POST':
        update_id = request.form['inputUpdateNumber']
        name = request.form['inputName']
        department = request.form['inputDepartment']
        phone = request.form['inputPhoneNumber']
        manager = request.form['inputManager']
        location = request.form['inputLocation']
        device = request.form.get('inputDevice')
        remarks = request.form.get('inputRemarks')
        object = request.form.get('inputObject')
        detail_location = request.form['inputDetailLocation']
        # print(update_id)

        update_visitor = Visitor.query.filter_by(id=update_id).first()
        update_visitor.name = name
        update_visitor.department = department
        update_visitor.phone = phone
        update_visitor.manager = manager
        if device == '1':
            update_visitor.device = True
        else:
            update_visitor.device = False
        update_visitor.remarks = remarks
        update_visitor.location = location
        update_visitor.detail_location = detail_location
        update_visitor.object = object
        db.session.commit()
        return redirect('manage_visitors')
    else:
        # 타임 스탬프
        today_weekday = datetime.now().weekday()
        weekdays = {0: "월요일", 1: "화요일", 2: "수요일", 3: "목요일", 4: "금요일", 5: "토요일", 6: "일요일"}
        weekday = weekdays.get(today_weekday, "")
        current_date = datetime.now().strftime('%Y년 %m월 %d일 ') + weekday
        current_time = datetime.now().strftime('\n%p %H:%M:%S')
        time = [ current_date, current_time]

        # 승인된 방문객 Sort_Desc
        approve_visitors = Visitor.query.filter_by(approve=1).order_by(Visitor.id.desc())

        # 출입 카드 목록
        card_list = Card.query.all()

        if approve_visitors:
            # 실시간 출입 방문객
            in_visitor = Visitor.query.filter_by(exit=0)
            in_visitor_card_none = Visitor.query.filter_by(exit=0, card_id=None)

            # 실시간 출입 방문객 카운팅
            in_visitor = in_visitor.count()
            in_visitor_card_none = in_visitor_card_none.count()
            in_visitor = in_visitor - in_visitor_card_none

            today = date.today()
            total_visitors = db.session.query(func.sum(Year.count)).scalar() # 총 방문객
            year = Year.query.filter_by(year=today.year).first() # 연간 방문객
            month = Month.query.filter_by(year=today.year, month=today.month).first() # 월간 방문객
            day = Day.query.filter_by(year=today.year, month=today.month, day=today.day).first() # 일간 방문객
            if year:
                yearly_visitor = year.count
                if month:
                    monthly_visitor = month.count
                    if day:
                        daily_visitor = day.count
                        visitor_count = [in_visitor, yearly_visitor, monthly_visitor, daily_visitor]
                    else:
                        visitor_count = [in_visitor, yearly_visitor, monthly_visitor, 0]
                else:
                    visitor_count = [in_visitor, yearly_visitor, 0, 0]
            else:
                visitor_count = [in_visitor, 0, 0, 0]

        # 내방객 등록 - 부서 목록
        department_lists = Department.query.all()
        return render_template('visitor_update.html', current_user=current_user, approve_visitors=approve_visitors, visitor_count=visitor_count, time=time, card_list=card_list, department_lists=department_lists, total_visitors=total_visitors)

#===================================================================================


#===================================================================================

# 부서 관리 페이지
@app.route('/departments', methods=['GET', 'POST'])
@login_required
def manage_departments():
    if request.method == 'POST':
        department_type = request.form['select_department_type']
        department_name = request.form['add_department_name_value']
        if Department.query.filter_by(department_name=department_name).first():
            flash('이미 부서가 존재합니다.')
            return redirect('departments')
        department = Department(department_type, department_name)
        db.session.add(department)
        db.session.commit()
        return redirect('departments')
    else:
        departments = Department.query.all()
        department_subsidiary = Department.query.filter_by(department_type="계열사").count()
        department_partner = Department.query.filter_by(department_type="협력사").count()
        department_bp = Department.query.filter_by(department_type="BP").count()
        department_etc = Department.query.filter_by(department_type="기타").count()
        department_counts = [department_subsidiary, department_partner, department_bp, department_etc]

        return render_template('manage_department.html', departments=departments, department_counts=department_counts)

# 부서 삭제 api
@app.route('/api/ajax_department_delete', methods=['POST'])
@login_required
def ajax_department_delete():
    data = request.get_json()
    print(data['delete_id'])
    delete_id = data['delete_id']
    department = Department.query.filter_by(id=delete_id).first()
    db.session.delete(department)
    db.session.commit()
    return jsonify()

# 기본 부서 생성
@app.route('/api/ajax_department_basic_create', methods=['POST'])
@login_required
def ajax_department_basic_create():
    # 출입 카드 DB Content 생성
    categories = ['CJ 올리브네트웍스', 'CJ(주)', 'CJ CGV', 'CJ 올리브영', 'CJ 프레시웨이', 'CJ 대한통운', 'CJ ENM', '디아이웨어']
    for category in categories:
        department = Department('계열사', category)
        db.session.add(department)
        db.session.commit()
    return jsonify(result = "success")

# 부서 초기화 api
@app.route('/api/ajax_department_reset', methods=['POST'])
@login_required
def ajax_department_reset():
    # 부서 테이블 초기화
    db.session.query(Department).delete()
    db.session.commit()
    return jsonify(result = "success")

#===================================================================================


#===================================================================================

# 방문객 수정 api
@app.route('/visit_update', methods=['GET', 'POST'])
@login_required
def visit_update():
    if request.method == 'POST':
        update_id = request.form['inputUpdateNumber']
        name = request.form['inputUpdateName']
        department = request.form['inputUpdateDepartment']
        phone = request.form['inputUpdatePhoneNumber']
        manager = request.form['inputUpdateManager']
        location = request.form['inputUpdateLocation']
        detail_location = request.form['inputUpdateDetailLocation']
        device = request.form.get('inputUpdateDevice')
        remarks = request.form.get('inputUpdateRemarks')
        object = request.form.get('inputUpdateObject')

        update_visitor = Visitor.query.filter_by(id=update_id).first()
        update_visitor.name = name
        update_visitor.department = department
        update_visitor.phone = phone
        update_visitor.manager = manager
        if device == '1':
            update_visitor.device = True
        else:
            update_visitor.device = False
        update_visitor.remarks = remarks
        update_visitor.location = location
        update_visitor.detail_location = detail_location
        update_visitor.object = object
        db.session.commit()
        return redirect('visit')

#===================================================================================


#===================================================================================

# 카드 관리 페이지
@app.route('/manage_cards', methods=['GET', 'POST'])
@login_required
def manage_cards():
    if request.method == 'POST':
        pass
    else:
        cards = db.session.query(Card.card_type).distinct().all()
        categories = []
        for card in cards:
            categories.append(card[0])

        card_counts = {}
        recall_cards = []
        use_cards = []
        for category in categories:
            card_counts[f'{category}'] = Card.query.filter_by(card_type=category).order_by(func.cast(Card.card_num, Integer).asc()).all()
            # 회수된 카드
            recall_cards.append(Card.query.filter_by(card_type=category, card_status='회수').count())

        # 총 출입 카드
        total_cards = Card.query.count()

        return render_template('manage_cards.html', card_counts=card_counts, recall_cards=recall_cards, total_cards=total_cards)

#===================================================================================


#===================================================================================

# 랙키 관리 페이지
@app.route('/manage_rack_keys', methods=['GET', 'POST'])
@login_required
def manage_rack_keys():
    if request.method == 'POST':
        pass
    else:
        keys = db.session.query(Rack.key_type).distinct().all()
        categories = []
        for key in keys:
            categories.append(key[0])

        key_counts = {}
        recall_keys = []
        use_keys = []
        for category in categories:
            key_counts[f'{category}'] = Rack.query.filter_by(key_type=category).order_by(func.cast(Rack.key_num, Integer).asc()).all()
            # 회수된 키
            recall_keys.append(Rack.query.filter_by(key_type=category, key_status='회수').count())

        # 총 출입 키
        total_keys = Rack.query.count()
        
        return render_template('manage_rack_key.html', key_counts=key_counts, recall_keys=recall_keys, total_keys=total_keys)

#===================================================================================


#===================================================================================

# 로그 관리 페이지
@app.route('/manage_logs', methods=['GET', 'POST'])
@login_required
def manage_logs():
    user_log = User_log.query.all()
    visitors = Visitor.query.all()
    return render_template('manage_logs.html', user_log=user_log)

@app.route('/<pagename>')
def admin(pagename):
    return render_template(pagename+'.html')

#===================================================================================


#===================================================================================

# 회원가입 페이지
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        # 폼으로부터 입력받은 데이터 가져오기
        username = request.form['username']
        email = request.form['email']
        password1 = request.form['password1']
        password2 = request.form['password2']
        department = request.form['registerDepartment']
        rank = request.form['registerRank']

        # 유효성 검사
        user = User.query.filter_by(email=email).first()
        if user:
            flash("이미 가입된 이메일입니다.")
        elif len(email) < 5 :
            flash("이메일은 5자 이상이어야 합니다.")
        elif len(email) > 30 :
            flash("이메일은 30자 이하여야 합니다.")
        elif len(username) < 2:
            flash("이름은 2자 이상이어야 합니다.")
        elif len(username) > 10:
            flash("이름은 10자 이하여야 합니다.")
        elif password1 != password2 :
            flash("비밀번호와 비밀번호재입력이 서로 다릅니다.")
        elif len(password1) < 7:
            flash("비밀번호는 8자 이상이어야 합니다.")
        elif len(password1) > 14:  # 비밀번호 길이 제한 추가
            flash("비밀번호는 14자 이하여야 합니다.")
        elif not re.search(r'[A-Z]', password1) or not re.search(r'[a-z]', password1) or not re.search(r'\d', password1) or not re.search(r'[!@#$%^&*(),.?":{}|<>]', password1):
            flash("비밀번호는 대문자, 숫자, 특수문자가 포함되어야 합니다.")
        else:
            # # 비밀번호 암호화
            hashed_password = bcrypt.generate_password_hash(password1)

            user = User(username, email, hashed_password, department, rank)
            db.session.add(user)
            db.session.commit()

            # 회원가입이 성공적으로 완료됨을 알리는 메시지 표시
            flash("회원가입 완료")
            return redirect('login')

    # GET 요청인 경우 회원가입 양식을 표시
    return render_template('register.html')

# 6개월 이상된 비밀번호 변경권고 코드 
def check_password_change(user):
    if user.password_changed_at:
        delta = datetime.now() - user.password_changed_at
        print("비밀번호 변경 시간:", user.password_changed_at)
        print("현재 시간:", datetime.now())
        if delta > timedelta(minutes=1):  # 6개월 이상 경과한 경우
            print("비밀번호 변경이 필요합니다.")
            print("현재 초과시간:", delta)
            flash("마지막 비밀번호 변경일로부터 6개월이 경과하였습니다. 비밀번호를 변경해주세요.")
            print("알림 메시지:", get_flashed_messages())

LOGIN_BLOCK_DURATION = 10 # 로그인 제한 기간 (분)
MAX_LOGIN_ATTEMPTS = 5 # 로그인 실패 허용 횟수


# 로그인 기능
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        current_date = datetime.now()
        email = request.form['email']
        password = request.form['password1']

        user = User.query.filter_by(email=email).first()
        if user:
            if user.login_blocked_until and user.login_blocked_until > datetime.now():
                # 로그인이 제한된 경우
                block_remaining = (user.login_blocked_until - datetime.now()).seconds // 60
                flash('일시적으로 로그인이 제한되었습니다. 잠시 후 다시 시도해주세요. (제한 시간: {}분 남음)'.format(block_remaining))
                return render_template('login.html')
            
            if bcrypt.check_password_hash(user.password, password):
                # 로그인 성공
                
                
                user.login_attempts = 0  # 로그인 시도 횟수 초기화
                user.login_blocked_until = None  # 로그인 제한 해제
                print('로그인 완료')
                
                #flash('마지막 비밀번호 변경일로부터 1분이 경과하였습니다. 비밀번호를 변경해주세요.')

                login_user(user, remember=True)
                user_log = User_log(request.remote_addr, current_date, user.id)
                
                
                db.session.add(user_log)
                db.session.commit()

                check_password_change(user) # 비밀번호 변경 주기 체크
                

                return redirect(url_for('index', user=user.username))
                #return render_template('index.html', user=user.username)

            
            else:
                # 로그인 실패
                user.login_attempts += 1  # 로그인 시도 횟수 증가
                if user.login_attempts >= MAX_LOGIN_ATTEMPTS:
                    # 일정 횟수 이상 실패 시 로그인 제한
                    user.login_blocked_until = datetime.now() + timedelta(minutes=LOGIN_BLOCK_DURATION)
                    block_remaining = LOGIN_BLOCK_DURATION
                    flash('로그인 횟수 제한까지 {}번 남았습니다. 잠시 후 다시 시도해주세요. (제한 시간: {}분)'.format(MAX_LOGIN_ATTEMPTS - user.login_attempts, block_remaining))
                else:
                    flash('비밀번호가 다릅니다. 로그인 횟수 제한까지 {}번 남았습니다.'.format(MAX_LOGIN_ATTEMPTS - user.login_attempts)) 
                # flash('비밀번호가 다릅니다.')
                db.session.commit()
                return render_template('login.html', login_attempts=user.login_attempts)
        else:
            flash('해당 이메일 정보가 없습니다.')
            return render_template('login.html')
    else: # GET 로그인 페이지
        return render_template('login.html')

# 로그아웃 페이지
@app.route('/logout')
@login_required
def logout():
    # 현재 시간
    current_date = datetime.now()
    # 현재 로그인된 사용자 중 가장 마지막 접속 기록 ID 추출
    logout_log = User_log.query.filter_by(user_id=current_user.id).order_by(User_log.id.desc()).first()
    # 마지막 접속 ID의 로그아웃 스탬프 컬럼에 현재 시간 기록
    logout_log.logout_timestamp = current_date
    # DB 적용
    db.session.commit()
    # 로그아웃 수행
    logout_user()
    print('logout success!')
    return redirect(url_for('login'))

#===================================================================================


#===================================================================================

# 내방객 등록 페이지
@app.route('/visit', methods=['GET', 'POST'])
@login_required
def visitor():
    if request.method == 'POST':
        name = request.form['inputName']
        department = request.form['inputDepartment']
        object = request.form['inputObject']
        phone = request.form['inputPhoneNumber']
        manager = request.form['inputManager']
        created_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        device = request.form.get('inputDevice')
        remarks = request.form.get('inputRemarks')
        work = request.form.get('inputWork')
        if device:
            device = True
            remarks = request.form.get('inputRemarks')
        else:
            device = False
            remarks = None
        if work:
            work = True
            location = request.form.get('inputLocation')
            detail_location = request.form.get('inputDetailLocation')
            company_type = request.form.get('inputCompany')
            company_name = request.form.get('inputCompanyName')
            work_content = request.form.get('inputContent')
        else:
            work = False
            location = None
            detail_location = None
            company_type = None
            company_name = None
            work_content = None

        # 내방객 등록하기 - 이름, 부서, 번호, 작업위치, 담당자, 장비체크, 비고, 방문목적, 등록시간, 승인, 사전/현장, 작업체크, 회사종류, 회사이름, 작업내용
        visitor = Visitor(name, department, phone, location, manager, device, remarks, object, created_time, 0, "사전 등록", work, company_type, company_name, work_content, detail_location)
        db.session.add(visitor)
        db.session.commit()
        return redirect(url_for('visitor'))
    else:
        # GET - 승인되지 않은 방문객 정보
        visitor_info = Visitor.query.filter_by(approve=0)

        # 내방객 등록 - 부서 목록
        department_lists = Department.query.all()
        print(department_lists)
        return render_template('visitor.html', department_lists=department_lists, visitor_info=visitor_info)


# 승인 버튼 클릭시 로직 ajax
@app.route('/api/ajax_approve', methods=['POST'])
@login_required
def ajax_approve():
    data = request.get_json()
    print(data['visitor_id'])
    print(data['approve'])

    visitor = Visitor.query.filter_by(id=data['visitor_id']).first()
    visitor.approve = 1
    visitor.exit = 0
    visitor.approve_date = datetime.now().strftime('%Y-%m-%d %H-%M-%S')
    visitor.approve_log = current_user.id

    today = date.today()
    year = Year.query.filter_by(year=today.year).first()
    if not year:
        year = Year(year=today.year, count=1)
        db.session.add(year)
    else:
        year.count += 1

    month = Month.query.filter_by(year=today.year, month=today.month).first()
    if not month:
        month = Month(year=today.year, month=today.month, count=1)
        db.session.add(month)
    else:
        month.count += 1

    day = Day.query.filter_by(year=today.year, month=today.month, day=today.day).first()
    if not day:
        day = Day(year=today.year, month=today.month, day=today.day, count=1)
        db.session.add(day)
    else:
        day.count += 1

    if visitor.device == 0:
        device = "미반입"
        privacy_device = False
    else:
        device = "반입"
        privacy_device = True

    if visitor.work == 0:
        work = "해당 없음"
        privacy_work = False
    else:
        work = "해당"
        privacy_work = True

    privacy_date = datetime.now().strftime('%Y-%m-%d %H-%M-%S')
    if visitor.registry == "사전 등록":
        image_send_sms_previous(visitor.name, visitor.approve_date, visitor.object, visitor.location, visitor.manager, visitor.phone, device, work, visitor.company_type, visitor.company, visitor.work_content)
        privacy = Privacy(visitor.name, visitor.department, visitor.phone, "", "", visitor.manager, privacy_device, privacy_work, visitor.remarks, visitor.object, visitor.location, visitor.detail_location, visitor.company_type, visitor.company, visitor.work_content, privacy_date, "사전 등록")
    else:
        image_send_sms_current(visitor.name, visitor.approve_date, visitor.object, visitor.location, visitor.manager, visitor.phone, device, work, visitor.company, visitor.work_content)
        privacy = Privacy(visitor.name, visitor.department, visitor.phone, "", "", visitor.manager, privacy_device, privacy_work, visitor.remarks, visitor.object, visitor.location, visitor.detail_location, "", visitor.company, visitor.work_content, privacy_date, "현장 등록")

    db.session.add(privacy)
    db.session.commit()
    return jsonify(result = "success")

# 반려 버튼 클릭시 로직 ajax
@app.route('/api/ajax_deny', methods=['POST'])
@login_required
def ajax_deny():
    data = request.get_json()
    print(data['visitor_id'])
    print(data['approve'])
    visitor = Visitor.query.filter_by(id=data['visitor_id']).first()
    db.session.delete(visitor)
    db.session.commit()
    return jsonify(result = "success")

# 긴급 승인 버튼 클릭시 로직 ajax
# @app.route('/api/ajax_emergency_approve', methods=['POST'])
# def ajax_emergency_approve():
#     data = request.get_json()
#     print(data['visitor_id'])
#     print(data['approve'])

#     visitor = Visitor.query.filter_by(id=data['visitor_id']).first()
#     visitor.approve = 1
#     visitor.exit = 0

#     today = date.today()
#     year = Year.query.filter_by(year=today.year).first()
#     if not year:
#         year = Year(year=today.year, count=1)
#         db.session.add(year)

#     month = Month.query.filter_by(year=today.year, month=today.month).first()
#     if not month:
#         month = Month(year=today.year, month=today.month, count=1)
#         db.session.add(month)

#     day = Day.query.filter_by(year=today.year, month=today.month, day=today.day).first()
#     if not day:
#         day = Day(year=today.year, month=today.month, day=today.day, count=1)
#         db.session.add(day)
#     else:
#         year.count += 1
#         month.count += 1
#         day.count += 1

#     db.session.commit()
#     return jsonify(result = "success")

#===================================================================================


#===================================================================================

# 내방객 관리 페이지 -  퇴실 버튼 클릭시 로직 ajax
@app.route('/api/ajax_exit', methods=['POST'])
@login_required
def ajax_exit():
    data = request.get_json()
    visitor = Visitor.query.filter_by(id=data['exit_id']).first()

    if visitor.card_id is None: # 카드를 안 받은 사람은 퇴실 X
        return "Card None"

    if visitor.exit_date == None and visitor.exit == 0:
        visitor.exit_date = datetime.now().strftime('%Y-%m-%d %H-%M-%S')
        visitor.exit = 1
        visitor.card.card_status = "회수"
        visitor.card_id = None
        if visitor.work == 1 and visitor.rack_id:
            visitor.rack.key_status = "회수"
            visitor.rack_id = None
        visitor.exit_log = current_user.id
        db.session.commit()
    else:
        return "Exit Error"

    return jsonify(response = "success")

# 내방객 관리 페이지 - 체크 박스 퇴실 api
@app.route('/api/ajax_index_exit_checkbox', methods=['POST'])
@login_required
def ajax_index_exit_checkbox():
    data = request.get_json()
    data_length = len(data['checked_datas'])

    if data_length < 1:
        return "No Select"

    for checked_data in data['checked_datas']:
        visitor = Visitor.query.filter_by(id=checked_data).first()
        if visitor.exit == 1:
            return "Exited"
        if visitor.card_id is None:
            return "No Card"

        if visitor.exit_date == None and visitor.exit == 0:
            visitor.exit_date = datetime.now().strftime('%Y-%m-%d %H-%M-%S')
            visitor.exit = 1
            visitor.card.card_status = "회수"
            visitor.card_id = None
            if visitor.work == 1 and visitor.rack_id:
                visitor.rack.key_status = "회수"
                visitor.rack_id = None
            visitor.exit_log = current_user.id
            db.session.commit()
        else:
            return "Exited"
    return jsonify(result = "success")

# 내방객 등록 페이지 - 체크 박스 승인 api
@app.route('/api/ajax_visit_approve_checkbox', methods=['POST'])
@login_required
def ajax_visit_approve_checkbox():
    data = request.get_json()
    data_length = len(data['checked_datas'])

    if data_length < 1:
        return "No Select"

    for checked_data in data['checked_datas']:
        visitor = Visitor.query.filter_by(id=checked_data).first()
        visitor.approve = 1
        visitor.exit = 0
        visitor.approve_date = datetime.now().strftime('%Y-%m-%d %H-%M-%S')
        visitor.approve_log = current_user.id

        today = date.today()
        year = Year.query.filter_by(year=today.year).first()
        if not year:
            year = Year(year=today.year, count=1)
            db.session.add(year)
        else:
            year.count += 1

        month = Month.query.filter_by(year=today.year, month=today.month).first()
        if not month:
            month = Month(year=today.year, month=today.month, count=1)
            db.session.add(month)
        else:
            month.count += 1

        day = Day.query.filter_by(year=today.year, month=today.month, day=today.day).first()
        if not day:
            day = Day(year=today.year, month=today.month, day=today.day, count=1)
            db.session.add(day)
        else:
            day.count += 1

        if visitor.device == 0:
            device = "미반입"
            privacy_device = False
        else:
            device = "반입"
            privacy_device = True

        if visitor.work == 0:
            work = "해당 없음"
            privacy_work = False
        else:
            work = "해당"
            privacy_work = True

        privacy_date = datetime.now().strftime('%Y-%m-%d %H-%M-%S')
        if visitor.registry == "사전 등록":
            image_send_sms_previous(visitor.name, visitor.approve_date, visitor.object, visitor.location, visitor.manager, visitor.phone, device, work, visitor.company_type, visitor.company, visitor.work_content)
            privacy = Privacy(visitor.name, visitor.department, visitor.phone, "", "", visitor.manager, privacy_device, privacy_work, visitor.remarks, visitor.object, visitor.location, visitor.detail_location, visitor.company_type, visitor.company, visitor.work_content, privacy_date, "사전 등록")
        else:
            image_send_sms_current(visitor.name, visitor.approve_date, visitor.object, visitor.location, visitor.manager, visitor.phone, device, work, visitor.company, visitor.work_content)
            privacy = Privacy(visitor.name, visitor.department, visitor.phone, "", "", visitor.manager, privacy_device, privacy_work, visitor.remarks, visitor.object, visitor.location, visitor.detail_location, "", visitor.company, visitor.work_content, privacy_date, "현장 등록")

        db.session.add(privacy)
        db.session.commit()
    return jsonify(result = "success")

# 내방객 등록 페이지 - 체크 박스 반려 api
@app.route('/api/ajax_visit_deny_checkbox', methods=['POST'])
@login_required
def ajax_visit_deny_checkbox():
    data = request.get_json()
    data_length = len(data['checked_datas'])

    if data_length < 1:
        return "No Select"

    for checked_data in data['checked_datas']:
        visitor = Visitor.query.filter_by(id=checked_data).first()
        db.session.delete(visitor)
        db.session.commit()
    return jsonify(result = "success")

# visit 체크 박스 긴급 승인 api - G6에게 이메일, 문자메세지
# @app.route('/api/ajax_visit_emergency_approve_checkbox', methods=['POST'])
# def ajax_visit_emergency_approve_checkbox():
#     data = request.get_json()
#     data_length = len(data['checked_datas'])

#     if data_length < 1:
#         return "No Select"

#     for checked_data in data['checked_datas']:
#         visitor = Visitor.query.filter_by(id=checked_data).first()
#         visitor.approve = 1
#         visitor.exit = 0

#         today = date.today()
#         year = Year.query.filter_by(year=today.year).first()
#         if not year:
#             year = Year(year=today.year)
#             db.session.add(year)

#         month = Month.query.filter_by(year=today.year, month=today.month).first()
#         if not month:
#             month = Month(year=today.year, month=today.month)
#             db.session.add(month)
#             year.count += 1

#         day = Day.query.filter_by(year=today.year, month=today.month, day=today.day).first()
#         if not day:
#             day = Day(year=today.year, month=today.month, day=today.day, count=1)
#             db.session.add(day)
#             month.count += 1
#         else:
#             year.count += 1
#             month.count += 1
#             day.count += 1

#         db.session.commit()
#     return jsonify(result = "success")

# 방문객 관리 페이지 - 담당자 업데이트 체크 박스 api
@app.route('/api/ajax_index_manager_update_checkbox', methods=['POST'])
@login_required
def ajax_index_manager_update_checkbox():
    data = request.get_json()
    manager = data['manager']
    data_length = len(data['checked_datas'])

    if data_length < 1:
        return "No Select"
    if manager is None or manager == "":
        return "Error"

    for checked_data in data['checked_datas']:
        visitor = Visitor.query.filter_by(id=checked_data).first()
        if visitor.exit == 1:
            return "Exited"
        visitor.manager = manager
        db.session.commit()
    return jsonify(result = "success")

# 방문객 관리 페이지 - 비고 업데이트 체크 박스 api
@app.route('/api/ajax_remarks_update_checkbox', methods=['POST'])
@login_required
def ajax_remarks_update_checkbox():
    data = request.get_json()
    remarks = data['remarks']
    data_length = len(data['checked_datas'])

    if data_length < 1:
        return "No Select"
    if remarks is None or remarks == "":
        return "Error"

    for checked_data in data['checked_datas']:
        visitor = Visitor.query.filter_by(id=checked_data).first()
        if visitor.exit == 1:
            return "Exited"
        visitor.remarks = remarks
        db.session.commit()
    return jsonify(result = "success")

# 상황실 방문객 관리 페이지 - 세부 작업 위치 업데이트 체크 박스 api
@app.route('/api/ajax_detail_location_update_checkbox', methods=['POST'])
@login_required
def ajax_detail_location_update_checkbox():
    data = request.get_json()
    detail_location = data['detail_location']
    data_length = len(data['checked_datas'])

    if data_length < 1:
        return "No Select"
    if detail_location is None or detail_location == "":
        return "Error"

    for checked_data in data['checked_datas']:
        visitor = Visitor.query.filter_by(id=checked_data).first()
        if visitor.exit == 1:
            return "Exited"
        visitor.detail_location = detail_location
        db.session.commit()
    return jsonify(result = "success")

# 내방객 관리 페이지 - 카드 불출 체크 박스 api
@app.route('/api/ajax_index_card_checkbox', methods=['POST'])
@login_required
def ajax_index_card_checkbox():
    data = request.get_json()
    card = data['card'].split(' ')
    print(card[0], card[1])
    data_length = len(data['checked_datas']) # 선택된 체크박스 수

    # 선택한 카드가 없거나 2개 이상 선택됐을 때 오류 발생
    if card is None:
        return "No Card"
    if data_length < 1:
        return "No Select"
    if data_length != 1:
        return "Multi Check"

    card_table = Card.query.filter_by(card_type=card[0], card_num=card[1]).first()
    for checked_data in data['checked_datas']:
        visitor = Visitor.query.filter_by(id=checked_data).first()
        if visitor.exit == 1:
            return "Exited"
        if visitor.card_id != None:
            return "Use Card"

        visitor.card_id = card_table.id
        card_table.card_status = "불출"
        db.session.commit()
    return jsonify(result = "success")

# 내방객 관리 페이지 - 카드 회수 체크 박스 api
@app.route('/api/ajax_update_visit_recall_card_checkbox', methods=['POST'])
@login_required
def ajax_update_visit_recall_card_checkbox():
    data = request.get_json()
    checked_datas = data['checked_datas']
    data_length = len(data['checked_datas']) # 선택된 체크박스 수
    print(checked_datas)

    # 선택한 카드가 없음
    if data_length < 1:
        return "No Select"
    
    for checked_data in checked_datas:
        recall_visitor = Visitor.query.filter_by(id=checked_data).first()
        if recall_visitor.card_id == None:
            return "No Card"
        if recall_visitor.exit == 1:
            return "Exited"
        recall_visitor.card.card_status = "회수"
        recall_visitor.card_id = None
        db.session.commit()
    return jsonify(result = "success")

# 내방객 등록, 내방객 관리 페이지 - 방문객 수정 api
@app.route('/api/ajax_update_manage_visit', methods=['POST'])
@login_required
def ajax_update_manage_visit():
    data = request.get_json()
    visitor = data['update_btn']
    update_visitor = Visitor.query.filter_by(id=visitor).first()
    if update_visitor.card_id != None:
        return "Use Card"
    else:
        update_visitor_info = [update_visitor.id, update_visitor.name, update_visitor.department, update_visitor.object, update_visitor.phone, update_visitor.manager, update_visitor.device, update_visitor.remarks, update_visitor.location, update_visitor.work, update_visitor.company_type, update_visitor.company, update_visitor.work_content, update_visitor.detail_location]
        return jsonify(response=update_visitor_info)


# 내방객 관리 페이지 - 방문객 삭제 api
@app.route('/api/ajax_delete_manage_visit', methods=['POST'])
@login_required
def ajax_delete_manage_visit():
    data = request.get_json()
    visitor = data['delete_btn']
    delete_visitor = Visitor.query.filter_by(id=visitor).first()
    if delete_visitor.card_id != None:
        return "Use Card"
    db.session.delete(delete_visitor)
    db.session.commit()
    return jsonify()


#===================================================================================


#===================================================================================

# MMS-IMAGE TEST SMS 문자 메세지 보내기 - 현장 등록 승인
def image_send_sms_current(name, date, object, location, manager, phone_num, device, work, company, work_content):
    connect_to_database()
    cursor = app.mysql_conn.cursor()  # 커서 생성
    insert_query = "INSERT INTO mms_msg (REQDATE, STATUS, TYPE, PHONE, CALLBACK, SUBJECT, MSG, FILE_CNT, FILE_TYPE1, FILE_PATH1) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
    if work == "해당":
        msg = name + '님 ' + '안녕하세요.\n' + '송도 IDC 센터에 방문하신 것을 환영합니다.\n-등록시간: ' + str(date) + '\n-방문위치: 인천광역시 연수구 하모니로177번길 20' + '\n-방문목적: ' + object + '\n-장비반입: ' + device + '\n-작업: ' + work + '\n-회사명: ' + company + '\n-작업내용:' + work_content + '\n-작업위치: ' + location + '\n-담당자: ' + manager + '\n-QR Code◀ '
    else:
        msg = name + '님 ' + '안녕하세요.\n' + '송도 IDC 센터에 방문하신 것을 환영합니다.\n-등록시간: ' + str(date) + '\n-방문위치: 인천광역시 연수구 하모니로177번길 20' + '\n-방문목적: ' + object + '\n-장비반입: ' + device + '\n-작업: ' + work + '\n-담당자: ' + manager + '\n-QR Code◀ '

    insert_data = (datetime.now(), '1', '0', phone_num, '01057320071', '[내방객 출입 관리 시스템 현장 등록 승인]', msg, '2', 'I', 'D://CJAgent//qr_img.jpg')  # 삽입할 데이터를 튜플로 정의
    cursor.execute(insert_query, insert_data)  # 쿼리 실행 및 데이터 전달
    app.mysql_conn.commit()  # 변경 사항 커밋
    cursor.close()  # 커서 닫기

# MMS-IMAGE TEST SMS 문자 메세지 보내기 - 사전 등록 승인
def image_send_sms_previous(name, date, object, location, manager, phone_num, device, work, company_type, company, work_content):
    connect_to_database()
    cursor = app.mysql_conn.cursor()  # 커서 생성
    insert_query = "INSERT INTO mms_msg (REQDATE, STATUS, TYPE, PHONE, CALLBACK, SUBJECT, MSG, FILE_CNT, FILE_TYPE1, FILE_PATH1) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
    if work == "해당":
        msg = name + '님 ' + '안녕하세요.\n' + '송도 IDC 센터에 방문하신 것을 환영합니다.\n-등록시간: ' + str(date) + '\n-방문위치: 인천광역시 연수구 하모니로177번길 20' + '\n-방문목적: ' + object + '\n-장비반입: ' + device + '\n-작업: ' + work + '\n-요청 회사: ' + company_type + '\n-회사명: ' + company + '\n-작업내용:' + work_content + '\n-작업위치: ' + location + '\n-담당자: ' + manager + '\n-QR Code◀ '
    else:
        msg = name + '님 ' + '안녕하세요.\n' + '송도 IDC 센터에 방문하신 것을 환영합니다.\n-등록시간: ' + str(date) + '\n-방문위치: 인천광역시 연수구 하모니로177번길 20' + '\n-방문목적: ' + object + '\n-장비반입: ' + device + '\n-작업: ' + work + '\n-담당자: ' + manager + '\n-QR Code◀ '

    insert_data = (datetime.now(), '1', '0', phone_num, '01057320071', '[내방객 출입 관리 시스템 사전 등록 승인]', msg, '2', 'I', 'D://CJAgent//qr_img.jpg')  # 삽입할 데이터를 튜플로 정의
    cursor.execute(insert_query, insert_data)  # 쿼리 실행 및 데이터 전달
    app.mysql_conn.commit()  # 변경 사항 커밋
    cursor.close()  # 커서 닫기

# manage qr 재전송 api
@app.route('/api/ajax_manage_qrcode_send', methods=['POST'])
@login_required
def ajax_manage_qrcode_send():
    data = request.get_json()
    qrcode_id = data['qrcode_btn']
    print(qrcode_id)
    qrcode_visitor = Visitor.query.filter_by(id=qrcode_id).first()
    print(qrcode_visitor.phone)
    if qrcode_visitor.card_id:
        return "Use Card"
    
    if qrcode_visitor.device == 0:
        device = "미반입"
    else:
        device = "반입"

    if qrcode_visitor.work == 0:
        work = "해당 없음"
    else:
        work = "해당"
    image_send_sms_previous(qrcode_visitor.name, qrcode_visitor.approve_date, qrcode_visitor.object, qrcode_visitor.location, qrcode_visitor.manager, qrcode_visitor.phone, device, work, qrcode_visitor.company_type, qrcode_visitor.company, qrcode_visitor.work_content)
    return jsonify()

#===================================================================================


#===================================================================================

# Excel 다운로드 api
@app.route('/api/ajax_excel_download', methods=['GET', 'POST'])
@login_required
def ajax_excel_download():
    data = request.get_json()
    start_date = data['start_date_val']
    end_date = data['end_date_val']

    if start_date == '' or end_date == '':
        return "No Date"
    if start_date > end_date:
        return "Date Error"

    # 내방객 점검 일지
    file_name = '일일 점검 양식.xlsx'
    
    # 파일 읽기
    workbook = openpyxl.load_workbook(file_name)  # 기존 파일을 로드합니다.
    sheet = workbook.active
    # 데이터 쓰기
    if start_date == end_date:
        # If start_date and end_date are the same, include only a single day
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        start_datetime_str = start_datetime.strftime('%Y-%m-%d')
        exited_visitors = Visitor.query.filter(
            and_(
                Visitor.approve_date >= start_datetime_str,
                Visitor.approve_date < (start_datetime + timedelta(days=1)).strftime('%Y-%m-%d'),
                Visitor.exit_date >= start_datetime_str,
                Visitor.exit_date < (start_datetime + timedelta(days=1)).strftime('%Y-%m-%d')
            )
        )
    else:
        # Include the range between start_date and end_date
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
        start_datetime_str = start_datetime.strftime('%Y-%m-%d')
        end_datetime_str = end_datetime.strftime('%Y-%m-%d')
        if start_datetime != end_datetime:
            end_datetime += timedelta(days=1)  # Add one day to include the end_date
        exited_visitors = Visitor.query.filter(
            and_(
                Visitor.approve_date >= start_datetime_str,
                Visitor.approve_date <= end_datetime.strftime('%Y-%m-%d'),
                Visitor.exit_date >= start_datetime_str,
                Visitor.exit_date <= end_datetime.strftime('%Y-%m-%d')
            )
        )

    for row_num, exited_visitor in enumerate(exited_visitors, 18):
        if exited_visitor.work == 1:
            work = 'O'
        else:
            work = 'X'

        if exited_visitor.device == 1:
            device = 'O'
        else:
            device = 'X'
        row = (
            exited_visitor.id, exited_visitor.approve_date, exited_visitor.exit_date,
            exited_visitor.department, exited_visitor.phone, exited_visitor.manager,
            work, exited_visitor.location, exited_visitor.detail_location,
            exited_visitor.company_type, exited_visitor.company, exited_visitor.work_content,
            device, "", exited_visitor.remarks
        )
        sheet.append(row)

        for col_num, _ in enumerate(row, 1):
            col_letter = get_column_letter(col_num)
            cell = sheet[col_letter + str(row_num)]
            cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet['A6'] = '일시: ' + start_date + ' ~ ' + end_date
    workbook.save('excel/' + start_date + '-' + end_date + ' 내방객 출입점검 일지1.xlsx')
    return jsonify(result="success")

#===================================================================================


#===================================================================================

# 카드 종류 추가 api
@app.route('/api/ajax_card_type_create', methods=['POST'])
@login_required
def ajax_card_type_create():
    data = request.get_json()
    print(data['card_type_val'])
    card_type_value = data['card_type_val']
    
    card_type_distinct = Card.query.filter_by(card_type=card_type_value).all()
    if card_type_distinct:
        return "type distinct"

    if card_type_value:
        card_type = Card(card_type_value, 1, '회수')
        db.session.add(card_type)
        db.session.commit()
    
    return jsonify(result="success")

# cards 카드 수량 추가 api
@app.route('/api/ajax_add_card', methods=['POST'])
@login_required
def ajax_add_card():
    data = request.get_json()
    card_number = data['add_card_value']
    card_type = data['select_card_type']
    print(card_type, card_number)

    if card_number == 0:
        return "No Number"
    
    cards = Card.query.filter_by(card_type=card_type).order_by(func.cast(Card.card_num, Integer).desc()).first()
    for index in range(1, int(card_number) + 1):
        add_cards = Card(card_type, int(cards.card_num)+int(index), '회수')
        db.session.add(add_cards)
        db.session.commit()
        
    print(cards.card_num)
    print(cards)
    return jsonify(result = "success")

# cards 카드 선택 추가 api
@app.route('/api/ajax_add_card_select', methods=['POST'])
@login_required
def ajax_add_card_select():
    data = request.get_json()
    card_number = data['add_card_value']
    card_type = data['select_card_type']
    print(card_type, card_number)
    
    card = Card.query.filter_by(card_type=card_type, card_num=card_number).first()
    if card != None:
        return "Use Card"
    
    add_card = Card(card_type, card_number, '회수')
    db.session.add(add_card)
    db.session.commit()

    return jsonify(result = "success")

# cards 카드 제거 api
@app.route('/api/ajax_card_delete', methods=['POST'])
@login_required
def ajax_card_delete():
    data = request.get_json()
    delete_id = data['delete_id']
    print(delete_id)

    delete_card = Card.query.filter_by(id=delete_id).first()
    db.session.delete(delete_card)
    db.session.commit()
    return jsonify(result = "success")

# 카드 분실 api
@app.route('/api/ajax_card_lose', methods=['POST'])
@login_required
def ajax_card_lose():
    data = request.get_json()
    card = data['card'].split(' ')
    card_lose = Card.query.filter_by(card_type=card[0], card_num=card[1]).first()
    card_lose.card_status = '분실'
    db.session.commit()
    
    return jsonify(result = "success")

# 카드 초기화 api
@app.route('/api/ajax_card_reset', methods=['POST'])
@login_required
def ajax_card_reset():
    # Card 테이블 초기화
    db.session.query(Card).delete()
    db.session.commit()
    return jsonify(result = "success")


#===================================================================================


#===================================================================================
# 키 종류 추가 api
@app.route('/api/ajax_key_type_create', methods=['POST'])
@login_required
def ajax_key_type_create():
    data = request.get_json()
    key_type_value = data['key_type_val']
    
    key_type_distinct = Rack.query.filter_by(key_type=key_type_value).all()
    if key_type_distinct:
        return "type distinct"

    if key_type_value:
        key_type = Rack(key_type_value, 1, '회수')
        db.session.add(key_type)
        db.session.commit()
    
    return jsonify(result="success")

# keys 키 수량 추가 api
@app.route('/api/ajax_add_key', methods=['POST'])
@login_required
def ajax_add_key():
    data = request.get_json()
    key_number = data['add_key_value']
    key_type = data['select_key_type']

    if key_number == 0:
        return "No Number"
    
    keys = Rack.query.filter_by(key_type=key_type).order_by(func.cast(Rack.key_num, Integer).desc()).first()
    for index in range(1, int(key_number) + 1):
        add_keys = Rack(key_type, int(keys.key_num)+int(index), '회수')
        db.session.add(add_keys)
        db.session.commit()
        
    return jsonify(result = "success")

# keys 키 선택 추가 api
@app.route('/api/ajax_add_key_select', methods=['POST'])
@login_required
def ajax_add_key_select():
    data = request.get_json()
    key_number = data['add_key_value']
    key_type = data['select_key_type']
    
    key = Rack.query.filter_by(key_type=key_type, key_num=key_number).first()
    if key != None:
        return "Use Key"
    
    add_key = Rack(key_type, key_number, '회수')
    db.session.add(add_key)
    db.session.commit()

    return jsonify(result = "success")

# 키 분실 api
@app.route('/api/ajax_key_lose', methods=['POST'])
@login_required
def ajax_key_lose():
    data = request.get_json()
    key = data['key'].split(' ')
    key_lose = Rack.query.filter_by(key_type=key[0], key_num=key[1]).first()
    key_lose.key_status = '분실'
    db.session.commit()
    
    return jsonify(result = "success")

# keys 키 제거 api
@app.route('/api/ajax_key_delete', methods=['POST'])
@login_required
def ajax_key_delete():
    data = request.get_json()
    delete_id = data['delete_id']

    delete_key = Rack.query.filter_by(id=delete_id).first()
    db.session.delete(delete_key)
    db.session.commit()
    return jsonify(result = "success")

# 키 초기화 api
@app.route('/api/ajax_key_reset', methods=['POST'])
@login_required
def ajax_key_reset():
    # 키 테이블 초기화
    db.session.query(Rack).delete()
    db.session.commit()
    return jsonify(result = "success")

#===================================================================================

# 유저 프로필 페이지
@app.route('/profile', methods=['GET', 'POST'])
@login_required
def user_profile():
    if request.method == 'POST':
        department = request.form['current_department']
        rank = request.form['current_rank']

        user = User.query.filter_by(email=current_user.email).first()
        if department and rank:
            user.department = department
            user.rank = rank
            db.session.commit()
            flash("프로필이 업데이트 되었습니다.")
            return redirect('profile')
        else:
            flash("프로필 업데이트에 실패했습니다.")
            return redirect('profile')
    return render_template('user_profile.html')

# 유저 인증 페이지
@app.route('/authenticated', methods=['GET', 'POST'])
@login_required
def user_authenticated():
    if request.method == 'POST':
        password = request.form['current_pwd']

        user = User.query.filter_by(email=current_user.email).first()
        if user:
            if bcrypt.check_password_hash(user.password, password):
                return redirect('profile_update')
            else:
                flash("비밀번호를 확인 후 다시 입력해주세요.")
    return render_template('authenticated.html')

# 유저 정보 변경 페이지
@app.route('/profile_update', methods=['GET', 'POST'])
@login_required
def user_profile_update():
    if request.method == 'POST':
        current_password = request.form['current_pwd']
        new_password_1 = request.form['new_pwd_1']
        new_password_2 = request.form['new_pwd_2']
        
        user = User.query.filter_by(email=current_user.email).first()
        
        if user:
            if bcrypt.check_password_hash(user.password, current_password):
                if current_password == new_password_1:
                    flash("현재 비밀번호와 새 비밀번호는 같을 수 없습니다.")
                elif new_password_1 != new_password_2:
                    flash("비밀번호와 비밀번호재입력이 서로 다릅니다.")
                elif len(new_password_1) < 7:
                    flash("비밀번호가 너무 짧습니다.")
                else:
                    hashed_password = bcrypt.generate_password_hash(new_password_1)
                    user.password = hashed_password
                    db.session.commit()
                    logout_user()
                    flash("비밀번호가 정상적으로 변경되었습니다.")
                    return redirect('login')
            else:
                flash("비밀번호를 잘못 입력하셨습니다.")

    return render_template('user_profile_update.html')

#===================================================================================

# 방문객 현장 등록 페이지
@app.route('/form-input', methods=['GET','POST'])
def form_input():
    if request.method == 'POST':
        date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # 필수 입력
        name = request.form['visitorName']
        department = request.form['visitorDepartment']
        phone = request.form['visitorPhone']
        birth = request.form['visitorBirth']
        manager = request.form['visitorManager']
        device = request.form['visitorDevice']
        remarks = request.form['visitorRemarks']
        object = request.form['visitorObject']

        # 선택 입력
        email = request.form.get('visitorEmail')
        location = request.form.get('visitorLocation')
        detail_location = request.form.get('visitorDetailLocation')
        company = request.form.get('visitorCompanyName')
        work_content = request.form.get('visitorWorkContent')
        if object == '작업':
            work = True
        else:
            work = False

        if device == '반입':
            device = True
        else:
            device = False

        privacy = Privacy(name, department, phone, birth, email, manager, device, work, remarks, object, location, detail_location, "", company, work_content, date, "현장 등록")
        visitor = Visitor(name, department, phone, location, manager, device, remarks, object, date, 0, "현장 등록", work, "", company, work_content, detail_location)
        db.session.add(privacy)
        db.session.add(visitor)
        db.session.commit()
        return redirect('form')
    return render_template('form-input.html')
# def __init__(self, name, department, phone, location, manager, device, remarks, object, created_time, approve, registry, work, company_type, company, work_content, detail_location):

#===================================================================================

@app.route('/rack_visitors')
@login_required
def manage_visitors_rack():
    # 타임 스탬프
    today_weekday = datetime.now().weekday()
    weekdays = {0: "월요일", 1: "화요일", 2: "수요일", 3: "목요일", 4: "금요일", 5: "토요일", 6: "일요일"}
    weekday = weekdays.get(today_weekday, "")
    current_date = datetime.now().strftime('%Y년 %m월 %d일 ') + weekday
    current_time = datetime.now().strftime('\n%p %H:%M:%S')
    time = [ current_date, current_time]

    # 승인된 방문객 Sort_Desc
    approve_visitors = Visitor.query.filter_by(approve=1).order_by(Visitor.id.desc())

    # 랙 키 목록
    rack_key_list = Rack.query.all()

    if approve_visitors:
        # 실시간 출입 방문객
        in_visitor = Visitor.query.filter_by(exit=0)
        in_visitor_card_none = Visitor.query.filter_by(exit=0, card_id=None)

        # 실시간 출입 방문객 카운팅
        in_visitor = in_visitor.count()
        in_visitor_card_none = in_visitor_card_none.count()
        in_visitor = in_visitor - in_visitor_card_none

        today = date.today()
        total_visitors = db.session.query(func.sum(Year.count)).scalar() # 총 방문객
        year = Year.query.filter_by(year=today.year).first() # 연간 방문객
        month = Month.query.filter_by(year=today.year, month=today.month).first() # 월간 방문객
        day = Day.query.filter_by(year=today.year, month=today.month, day=today.day).first() # 일간 방문객
        if year:
            yearly_visitor = year.count
            if month:
                monthly_visitor = month.count
                if day:
                    daily_visitor = day.count
                    visitor_count = [in_visitor, yearly_visitor, monthly_visitor, daily_visitor]
                else:
                    visitor_count = [in_visitor, yearly_visitor, monthly_visitor, 0]
            else:
                visitor_count = [in_visitor, yearly_visitor, 0, 0]
        else:
            visitor_count = [in_visitor, 0, 0, 0]
    
    # print('현재 로그인한 사용자: ' + str(current_user))
    return render_template('visitor_emergency.html', current_user=current_user, approve_visitors=approve_visitors, visitor_count=visitor_count, time=time, rack_key_list=rack_key_list, total_visitors=total_visitors)

# 키 불출 체크 박스 api
@app.route('/api/ajax_use_rack_key_checkbox', methods=['POST'])
@login_required
def ajax_use_rack_key_checkbox():
    data = request.get_json()
    key = data['key'].split(' ')
    print(key[0], key[1])
    data_length = len(data['checked_datas']) # 선택된 체크박스 수

    # 선택한 키가 없을 때 선택하지 않았을 때 오류 발생
    if key is None:
        return "No Key"
    if data_length < 1:
        return "No Select"

    key_table = Rack.query.filter_by(key_type=key[0], key_num=key[1]).first()
    for checked_data in data['checked_datas']:
        visitor = Visitor.query.filter_by(id=checked_data).first()
        if visitor.card_id == None:
            return "No Card"
        if visitor.exit == 1:
            return "Exited"
        if visitor.rack_id != None:
            return "Use Key"
        if visitor.work == False:
            return "No Work"

        visitor.rack_id = key_table.id
        key_table.key_status = "불출"
        db.session.commit()
    return jsonify(result = "success")

# 키 회수 체크 박스 api
@app.route('/api/ajax_recall_rack_key_checkbox', methods=['POST'])
@login_required
def ajax_recall_rack_key_checkbox():
    data = request.get_json()
    checked_datas = data['checked_datas']
    data_length = len(data['checked_datas']) # 선택된 체크박스 수

    # 선택한 키가 없음
    if data_length < 1:
        return "No Select"
    
    for checked_data in checked_datas:
        recall_visitor = Visitor.query.filter_by(id=checked_data).first()
        if recall_visitor.rack_id == None:
            return "No Key"
        if recall_visitor.exit == 1:
            return "Exited"
        recall_visitor.rack.key_status = "회수"
        recall_visitor.rack_id = None
        db.session.commit()
    return jsonify(result = "success")

#===================================================================================





@app.errorhandler(jinja2.exceptions.TemplateNotFound)
def template_not_found(e):
    return not_found(e)

@app.errorhandler(404)
def not_found(e):
    return render_template('404.html')

# MySQL 데이터베이스에 연결하는 함수
def connect_to_database():
    db_user, db_password, db_host, db_port, db_name = db_connector()
    app.config['DB_HOST'] = db_host
    app.config['DB_USER'] = db_user
    app.config['DB_PASSWORD'] = db_password
    app.config['DB_DATABASE'] = db_name
    app.config['DB_PORT'] = db_port
    app.mysql_conn = mysql.connector.connect(
        host=app.config['DB_HOST'],
        user=app.config['DB_USER'],
        password=app.config['DB_PASSWORD'],
        database=app.config['DB_DATABASE'],
        port=app.config['DB_PORT']
    )

if __name__ == '__main__':
    connect_to_database()
    # csrf.init_app(app) # CSRF Config
    app.run(debug=True)
